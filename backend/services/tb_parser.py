"""
Trial Balance Parser — handles the Dubai Women Establishment TB format.

Key facts about the actual file:
- Single sheet: GRP_Trial_Balance_report_with__
- 5 metadata rows above column headers (row 6 = headers, rows 7+ = data)
- No separate Debit/Credit columns — uses signed ending balances
- Account code format: NNNNN-NN
- Account Type column provides initial classification hint
- NaN rows represent dormant chart-of-account entries (kept but flagged)
"""

import re
from pathlib import Path
from typing import Optional
import pandas as pd
import openpyxl


# Column name patterns to detect TB columns regardless of exact header text
_CODE_PATTERNS = [
    r"account.*sub.*account", r"account.*code", r"account.*no", r"code",
    r"ledger.*code", r"gl.*code"
]
_NAME_PATTERNS = [
    r"account.*description", r"description", r"account.*name", r"name"
]
_SUB_PATTERNS = [
    r"sub.*account.*description", r"sub.*description", r"sub.*name", r"cost.*center"
]
_TYPE_PATTERNS = [
    r"account.*type", r"type", r"classification", r"category"
]
_BEG_PATTERNS = [
    r"beginning.*balance", r"opening.*balance", r"prior.*year", r"prev.*year",
    r"balance.*brought", r"b/f"
]
_ACTIVITY_PATTERNS = [
    r"period.*activity", r"movement", r"net.*movement", r"ytd.*movement",
    r"debit.*credit", r"net.*change"
]
_END_PATTERNS = [
    r"ending.*balance", r"closing.*balance", r"balance.*c/f",
    r"current.*year", r"year.*end", r"^balance$"
]
_DEBIT_PATTERNS = [r"^debit$", r"^dr\.?$", r"total.*debit"]
_CREDIT_PATTERNS = [r"^credit$", r"^cr\.?$", r"total.*credit"]


def _match_col(header: str, patterns: list[str]) -> bool:
    h = str(header).lower().strip()
    return any(re.search(p, h) for p in patterns)


def _detect_header_row(ws) -> int:
    """Find which row contains the column headers (look for typical TB header words)."""
    keywords = {"account", "description", "balance", "debit", "credit",
                "code", "type", "activity", "name"}
    for i, row in enumerate(ws.iter_rows(max_row=20, values_only=True), start=1):
        non_null = [v for v in row if v is not None]
        # Header rows have multiple non-null cells and 3+ keyword matches
        if len(non_null) < 3:
            continue
        row_text = " ".join(str(v).lower() for v in non_null)
        matches = sum(1 for kw in keywords if kw in row_text)
        if matches >= 3:
            return i
    return 1


def parse_tb(file_path: str | Path) -> dict:
    """
    Parse a Trial Balance Excel file.

    Returns a dict with:
      - accounts: list of dicts (one per TB row)
      - columns: detected column mapping
      - totals: summary totals
      - validation: basic checks
      - metadata: entity/period info
      - issues: list of structural issues
    """
    path = Path(file_path)
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb.active  # first/only sheet

    # ── Metadata from top rows ───────────────────────────────────────────────
    metadata = {}
    for row in ws.iter_rows(max_row=6, values_only=True):
        for cell in row:
            if cell is None:
                continue
            text = str(cell).strip()
            if "ledger name" in text.lower():
                metadata["entity_name"] = text.split(":", 1)[-1].strip()
            elif "period name" in text.lower():
                metadata["period"] = text.split(":", 1)[-1].strip()
            elif "amount type" in text.lower():
                metadata["amount_type"] = text.split(":", 1)[-1].strip()

    # ── Detect header row ────────────────────────────────────────────────────
    header_row = _detect_header_row(ws)
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]

    # ── Map columns by header name ───────────────────────────────────────────
    col_map = {}
    for idx, h in enumerate(headers):
        if h is None:
            continue
        if _match_col(h, _CODE_PATTERNS) and "code" not in col_map:
            col_map["code"] = idx
        if _match_col(h, _NAME_PATTERNS) and "name" not in col_map:
            col_map["name"] = idx
        if _match_col(h, _SUB_PATTERNS) and "sub" not in col_map:
            col_map["sub"] = idx
        if _match_col(h, _TYPE_PATTERNS) and "type" not in col_map:
            col_map["type"] = idx
        if _match_col(h, _BEG_PATTERNS) and "beginning" not in col_map:
            col_map["beginning"] = idx
        if _match_col(h, _ACTIVITY_PATTERNS) and "activity" not in col_map:
            col_map["activity"] = idx
        if _match_col(h, _END_PATTERNS) and "ending" not in col_map:
            col_map["ending"] = idx
        if _match_col(h, _DEBIT_PATTERNS) and "debit" not in col_map:
            col_map["debit"] = idx
        if _match_col(h, _CREDIT_PATTERNS) and "credit" not in col_map:
            col_map["credit"] = idx

    # ── Load data with pandas ────────────────────────────────────────────────
    df = pd.read_excel(str(path), sheet_name=0, header=header_row - 1, dtype=str)
    df = df.dropna(how="all")

    accounts = []
    issues = []
    seen_codes = {}

    for row_idx, row in df.iterrows():
        vals = row.tolist()

        def get(key) -> Optional[str]:
            if key in col_map and col_map[key] < len(vals):
                v = vals[col_map[key]]
                return str(v).strip() if pd.notna(v) and str(v).strip() not in ("nan", "") else None
            return None

        def get_num(key) -> float:
            raw = get(key)
            if raw is None:
                return 0.0
            try:
                return float(str(raw).replace(",", "").replace(" ", ""))
            except ValueError:
                return 0.0

        code = get("code")
        name = get("name")

        # Skip rows that aren't real accounts (summary/total rows, header repeats)
        if code is None and name is None:
            continue
        if code and not re.match(r"^\d{4,6}", code):
            continue

        sub = get("sub")
        acc_type = get("type")
        beginning = get_num("beginning")
        activity = get_num("activity")
        ending = get_num("ending")

        # Handle separate Debit/Credit format if present
        if "debit" in col_map and "credit" in col_map:
            dr = get_num("debit")
            cr = get_num("credit")
            ending = dr - cr
            beginning = 0.0
            activity = ending

        # Arithmetic check
        calc_ending = beginning + activity
        arith_ok = abs(calc_ending - ending) < 0.01 if (beginning != 0 or activity != 0) else True

        # Anomaly flags
        is_zero = abs(ending) < 0.001 and abs(beginning) < 0.001 and abs(activity) < 0.001
        is_unusual = False
        unusual_reason = None

        if acc_type:
            at_lower = acc_type.lower()
            # Expense with credit balance
            if "expense" in at_lower and ending < -0.01:
                is_unusual = True
                unusual_reason = f"Expense account with credit balance: {ending:,.2f}"
            # Liability/equity with debit balance (positive = debit for these)
            elif "liability" in at_lower and ending > 0.01:
                is_unusual = True
                unusual_reason = f"Liability account with debit balance: {ending:,.2f}"
            elif "revenue" in at_lower and ending > 0.01:
                is_unusual = True
                unusual_reason = f"Revenue account with debit balance: {ending:,.2f}"

        # Duplicate detection
        if code:
            if code in seen_codes:
                issues.append({
                    "type": "DUPLICATE_CODE",
                    "code": code,
                    "rows": [seen_codes[code], int(row_idx) + 1]
                })
            else:
                seen_codes[code] = int(row_idx) + 1

        accounts.append({
            "account_code": code or "",
            "account_name": name or "(unnamed)",
            "sub_account": sub,
            "account_type_raw": acc_type or "Unknown",
            "beginning_balance": beginning,
            "period_activity": activity,
            "ending_balance": ending,
            "source_row": int(row_idx) + 2,  # 1-based Excel row
            "is_zero": is_zero,
            "is_unusual": is_unusual,
            "unusual_reason": unusual_reason,
            "arith_ok": arith_ok,
        })

    # ── Totals ───────────────────────────────────────────────────────────────
    total_ending = sum(a["ending_balance"] for a in accounts)
    total_beginning = sum(a["beginning_balance"] for a in accounts)
    total_activity = sum(a["period_activity"] for a in accounts)

    # By type
    by_type: dict[str, float] = {}
    for a in accounts:
        t = a["account_type_raw"]
        by_type[t] = by_type.get(t, 0.0) + a["ending_balance"]

    # ── Validation checks ────────────────────────────────────────────────────
    validation = []

    balanced = abs(total_ending) < 1.0
    validation.append({
        "check": "TB Balance (net to zero)",
        "result": "PASS" if balanced else "ERROR",
        "expected": "0.00",
        "actual": f"{total_ending:,.2f}",
        "difference": f"{total_ending:,.2f}",
        "severity": "CRITICAL" if not balanced else "OK",
        "explanation": "Total ending balance should net to zero in a balanced TB." if not balanced
                       else "Trial balance is balanced.",
    })

    zero_count = sum(1 for a in accounts if a["is_zero"])
    validation.append({
        "check": "Zero-balance accounts",
        "result": "WARNING" if zero_count > 0 else "PASS",
        "expected": "0",
        "actual": str(zero_count),
        "difference": str(zero_count),
        "severity": "LOW" if zero_count > 0 else "OK",
        "explanation": f"{zero_count} account(s) have zero balance across all periods.",
    })

    unusual_count = sum(1 for a in accounts if a["is_unusual"])
    validation.append({
        "check": "Unusual debit/credit balances",
        "result": "WARNING" if unusual_count > 0 else "PASS",
        "expected": "0",
        "actual": str(unusual_count),
        "difference": str(unusual_count),
        "severity": "MEDIUM" if unusual_count > 0 else "OK",
        "explanation": f"{unusual_count} account(s) have unexpected debit/credit balances for their type.",
    })

    arith_fails = [a for a in accounts if not a.get("arith_ok", True)]
    validation.append({
        "check": "Arithmetic: Beginning + Activity = Ending",
        "result": "WARNING" if arith_fails else "PASS",
        "expected": "0 failures",
        "actual": f"{len(arith_fails)} failures",
        "difference": str(len(arith_fails)),
        "severity": "MEDIUM" if arith_fails else "OK",
        "explanation": f"{len(arith_fails)} account(s) fail the Beg + Activity = Ending check.",
    })

    if issues:
        dup_count = len([i for i in issues if i["type"] == "DUPLICATE_CODE"])
        validation.append({
            "check": "Duplicate account codes",
            "result": "ERROR",
            "expected": "0",
            "actual": str(dup_count),
            "difference": str(dup_count),
            "severity": "HIGH",
            "explanation": f"{dup_count} duplicate account code(s) found.",
        })

    return {
        "accounts": accounts,
        "columns": col_map,
        "headers": headers,
        "metadata": metadata,
        "totals": {
            "total_ending": total_ending,
            "total_beginning": total_beginning,
            "total_activity": total_activity,
            "by_type": by_type,
            "account_count": len(accounts),
            "zero_count": zero_count,
            "unusual_count": unusual_count,
        },
        "validation": validation,
        "issues": issues,
        "balanced": balanced,
        "sheet_name": ws.title,
    }
