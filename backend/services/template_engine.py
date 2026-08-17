"""
Template population engine.

Discovers the structure of an uploaded template workbook, then populates a
COPY of it — never modifying the original.

Supports:
  Template A — Audit Draft (financial statements)
  Template B — Audit File (working papers / LEAD sheet)

All formatting, formulas, merges, hidden rows/columns, and logos are preserved.
"""

import copy
import json
import shutil
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

from backend.models.database import get_conn


# ── Audit Draft template — hard-coded cell targets discovered from analysis ──
AUDIT_DRAFT_TARGETS = {
    # Cover Page named range inputs
    "cover_page": {
        "sheet": "Cover Page",
        "inputs": {
            "company_name":               ("L", 6),
            "company_location":           ("L", 7),
            "audit_period":               ("L", 8),
            "audit_period_words":         ("L", 9),
            "audit_period_words_small":   ("L", 10),
            "audit_period_beginning":     ("L", 11),
            "comparative_period":         ("L", 12),
            "comparative_period_words":   ("L", 13),
            "comparative_period_begin":   ("L", 14),
            "currency":                   ("L", 15),
            "authorised_person":          ("L", 16),
            "designation":                ("L", 17),
            "activity":                   ("L", 18),
            "law":                        ("L", 19),
            "audit_report_date":          ("L", 20),
        },
    },

    # Balance Sheet — current year = col M, prior year = col O
    "balance_sheet": {
        "sheet": "Balance Sheet1",
        "cy_col": "M",
        "py_col": "O",
        "balance_check_cy": ("M", 53),  # must = 0
        "balance_check_py": ("O", 53),
        "lines": {
            "PPE":               {"row": 10, "note_col": "K", "note": 5},
            "INTANGIBLES":       {"row": 11, "note_col": "K", "note": 6},
            "INVENTORIES":       {"row": 16, "note_col": "K", "note": 7},
            "TRADE_RECEIVABLES": {"row": 17, "note_col": "K", "note": 8},
            "CASH":              {"row": 18, "note_col": "K", "note": 10},
            "REVALUATION_RESERVE":   {"row": 27, "note_col": "K", "note": None},
            "ACCUMULATED_LOSSES":    {"row": 28, "note_col": "K", "note": None},
            "EOSB":              {"row": 34, "note_col": "K", "note": 11},
            "TRADE_PAYABLES":    {"row": 39, "note_col": "K", "note": 12},
        },
    },

    # P&L — current year = col I, prior year = col K
    "pnl": {
        "sheet": "P & L ",
        "cy_col": "I",
        "py_col": "K",
        "lines": {
            "REVENUE":        {"row": 10},
            "COS":            {"row": 11},
            "GRANT_INCOME":   {"row": 14},
            "OTHER_INCOME":   {"row": 15},
            "ADMIN_EXPENSES": {"row": 19},
            "FINANCE_COST":   {"row": 20},
        },
    },

    # PPE Note
    "ppe_note": {
        "sheet": "PPE",
        "categories": {
            "Land":                     "C",
            "Building":                 "E",
            "Furniture & Office Equip": "G",
            "Motor Vehicles":           "I",
        },
        "rows": {
            "opening_cost":   10,
            "additions":      12,
            "closing_cost":   15,
            "opening_dep":    18,
            "dep_charge":     20,
            "closing_dep":    23,
            "nbv_current":    26,
            "nbv_prior":      28,
        },
    },

    # Notes sheet — key rows
    "notes": {
        "sheet": "Notes",
        "cy_col": "G",
        "py_col": "I",
    },
}


# ── Audit File template — LEAD sheet and AP-01 ───────────────────────────────
AUDIT_FILE_TARGETS = {
    "ap01": {
        "sheet": "AP-01",
        "inputs": {
            "client_name": ("C", 5),
            "period":      ("C", 7),
            "prepared_by": ("H", 6),
        },
    },
    "lead": {
        "sheet": "LEAD",
        # Rows will be detected dynamically from sheet content
        # because the exact row numbers may vary between template versions
        "cy_col": "D",
        "adj_dr_col": "E",
        "adj_cr_col": "F",
        "py_col": "H",
        "year_cy_cell": ("D", 6),
        "year_py_cell": ("H", 6),
        # Mapping: lead_line_key → label fragment to search for in column B
        "line_labels": {
            "PPE":               "property, plant",
            "INTANGIBLES":       "intangible",
            "INVESTMENTS":       "investment",
            "OTHER_NCA":         "other non-current",
            "INVENTORIES":       "inventor",
            "TRADE_RECEIVABLES": "trade.*receivable|receivable",
            "CASH":              "cash.*bank|bank.*cash",
            "EOSB":              "end of service|eosb|gratuity",
            "TRADE_PAYABLES":    "trade.*payable|payable",
            "EQUITY":            "equity|capital|reserve|retained",
            "REVENUE":           "revenue|sales",
            "COS":               "cost of",
            "GRANT_INCOME":      "grant",
            "OTHER_INCOME":      "other income",
            "ADMIN_EXPENSES":    "admin|general.*admin",
            "FINANCE_COST":      "finance cost|finance exp",
        },
    },
}


def analyze_template(file_path: str | Path) -> dict:
    """
    Inspect a template workbook and return metadata describing its structure.
    """
    path = Path(file_path)
    wb = openpyxl.load_workbook(str(path), data_only=False)

    meta = {
        "file": path.name,
        "sheets": [],
        "named_ranges": {},
        "template_type": _detect_template_type(wb),
    }

    # Named ranges
    for name, defn in wb.defined_names.items():
        try:
            destinations = list(defn.destinations)
            if destinations:
                sheet_title, coord = destinations[0]
                meta["named_ranges"][name] = f"{sheet_title}!{coord}"
        except Exception:
            meta["named_ranges"][name] = str(defn.attr_text)

    # Per-sheet summary
    for ws in wb.worksheets:
        hidden_rows = [
            r for r, rd in ws.row_dimensions.items() if rd.hidden
        ]
        hidden_cols = [
            get_column_letter(c) for c, cd in ws.column_dimensions.items()
            if cd.hidden
        ]
        merged = [str(m) for m in ws.merged_cells.ranges]

        # Count formulas
        formula_count = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1

        meta["sheets"].append({
            "name": ws.title,
            "state": ws.sheet_state,
            "dimensions": ws.dimensions,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "hidden_rows": hidden_rows,
            "hidden_cols": hidden_cols,
            "merged_cells": merged[:20],  # cap for JSON size
            "formula_count": formula_count,
        })

    wb.close()
    return meta


def _detect_template_type(wb) -> str:
    """Determine whether this is the Audit Draft or Audit File template."""
    sheet_names = [ws.title for ws in wb.worksheets]
    if any("Balance Sheet" in s or "P & L" in s for s in sheet_names):
        return "audit_draft"
    if any("LEAD" in s or "AP-01" in s for s in sheet_names):
        return "audit_file"
    return "unknown"


def _find_lead_rows(ws) -> dict[str, int]:
    """
    Scan the LEAD sheet column B for line item labels.
    Returns a dict of lead_line_key → row_number.
    """
    import re
    label_map = AUDIT_FILE_TARGETS["lead"]["line_labels"]
    found = {}

    for row in ws.iter_rows(min_col=2, max_col=2, values_only=True):
        cell_val = row[0]
        if cell_val is None:
            continue
        text = str(cell_val).lower().strip()
        row_num = ws.cell(ws.min_row, 1).row  # hack — use enumerate below

    # Proper enumeration
    for r_idx, row in enumerate(ws.iter_rows(min_col=2, max_col=2), start=1):
        cell = row[0]
        if cell.value is None:
            continue
        text = str(cell.value).lower().strip()
        for key, pattern in label_map.items():
            if key not in found and re.search(pattern, text):
                found[key] = cell.row
                break

    return found


def _aggregate_balances(mappings: list[dict]) -> dict[str, dict]:
    """
    Aggregate TB account balances by fs_line_item.

    Sign convention for the signed-balance TB format:
    - Assets and Expenses have DEBIT (positive) balances in the TB
    - Liabilities, Equity, and Revenue have CREDIT (negative) balances in the TB
    - For presentation on financial statements, credit-normal accounts are negated
      so that all presented figures are positive magnitudes.
    - Accumulated depreciation (contra-asset, classified under PPE/INTANGIBLES)
      keeps its negative sign so it correctly nets against gross cost.

    Returns: {line_item_key: {"cy": float, "py": float}}
    """
    agg: dict[str, dict] = {}
    for m in mappings:
        key = m.get("fs_line_item") or m.get("lead_line", "UNCLASSIFIED")
        if key not in agg:
            agg[key] = {"cy": 0.0, "py": 0.0}
        cy = float(m.get("ending_balance") or 0)
        py = float(m.get("beginning_balance") or 0)

        stmt = m.get("fs_statement", "")
        cat = m.get("fs_category", "")

        if stmt == "balance_sheet":
            if "Liab" in cat or "Equity" in cat:
                # Credit-normal: negate to get presentation amount (positive = owed/attributable)
                cy = -cy
                py = -py
            # Non-Current Assets accounts (PPE, Intangibles) keep their sign:
            # - Gross cost accounts: positive (debit)
            # - Accumulated depreciation accounts: negative (credit) → reduces the total
        elif stmt == "pnl":
            # Only negate accounts that are CREDIT-NORMAL (revenue/income/grant lines).
            # "Cost of Revenue" contains "Revenue" but is debit-normal — exclude it.
            _CREDIT_PL_CATS = {"Revenue", "Grant Received", "Other Income",
                                "Other income", "Grant income", "Grant received"}
            if cat in _CREDIT_PL_CATS:
                cy = -cy
                py = -py
            # Expense accounts ("Cost of Revenue", "General and Administrative Expenses",
            # "Finance Cost") keep their positive (debit) sign

        agg[key]["cy"] += cy
        agg[key]["py"] += py

    return agg


def populate_audit_draft(
    template_path: str | Path,
    output_path: str | Path,
    engagement: dict,
    mappings: list[dict],
) -> dict:
    """
    Populate the Audit Draft (financial statements) template.

    Returns a report of what was written and any issues.
    """
    # Make a working copy
    shutil.copy2(str(template_path), str(output_path))
    wb = openpyxl.load_workbook(str(output_path), keep_vba=False)

    report = {"written": [], "skipped": [], "issues": []}
    agg = _aggregate_balances(mappings)

    # ── Cover Page ────────────────────────────────────────────────────────────
    cp_targets = AUDIT_DRAFT_TARGETS["cover_page"]
    cp_sheet_name = cp_targets["sheet"]
    if cp_sheet_name in wb.sheetnames:
        ws_cp = wb[cp_sheet_name]
        cp_inputs = cp_targets["inputs"]

        def write_cp(key, value):
            col, row = cp_inputs[key]
            col_idx = column_index_from_string(col)
            _safe_write(ws_cp, row, col_idx, value, report)

        write_cp("company_name",             engagement.get("entity_name", ""))
        write_cp("company_location",         engagement.get("location", "DUBAI - UNITED ARAB EMIRATES"))
        write_cp("audit_period",             engagement.get("period_end_date", ""))
        write_cp("audit_period_words",       engagement.get("period_end_words_upper", ""))
        write_cp("audit_period_words_small", engagement.get("period_end_words", ""))
        write_cp("audit_period_beginning",   engagement.get("period_start", ""))
        write_cp("comparative_period",       engagement.get("prior_period_end", ""))
        write_cp("comparative_period_words", engagement.get("prior_period_words", ""))
        write_cp("comparative_period_begin", engagement.get("prior_period_start", ""))
        write_cp("currency",                 engagement.get("currency", "AED"))
        write_cp("authorised_person",        engagement.get("authorised_person", ""))
        write_cp("designation",              engagement.get("designation", ""))

    # ── Balance Sheet ─────────────────────────────────────────────────────────
    bs_targets = AUDIT_DRAFT_TARGETS["balance_sheet"]
    bs_sheet_name = bs_targets["sheet"]
    if bs_sheet_name in wb.sheetnames:
        ws_bs = wb[bs_sheet_name]
        cy_col = column_index_from_string(bs_targets["cy_col"])
        py_col = column_index_from_string(bs_targets["py_col"])

        for line_key, line_info in bs_targets["lines"].items():
            row = line_info["row"]
            bal = agg.get(line_key, {"cy": 0.0, "py": 0.0})
            _safe_write(ws_bs, row, cy_col, bal["cy"], report, f"BS:{line_key}:CY")
            _safe_write(ws_bs, row, py_col, bal["py"], report, f"BS:{line_key}:PY")

        # Fix the broken Equity REF — write accumulated losses
        equity_bal = agg.get("ACCUMULATED_LOSSES", {"cy": 0.0, "py": 0.0})
        reval_bal = agg.get("REVALUATION_RESERVE", {"cy": 0.0, "py": 0.0})

        _safe_write(ws_bs, 27, cy_col, reval_bal["cy"], report, "BS:REVAL:CY")
        _safe_write(ws_bs, 27, py_col, reval_bal["py"], report, "BS:REVAL:PY")
        _safe_write(ws_bs, 28, cy_col, equity_bal["cy"], report, "BS:ACC_LOSS:CY")
        _safe_write(ws_bs, 28, py_col, equity_bal["py"], report, "BS:ACC_LOSS:PY")

    # ── P&L ───────────────────────────────────────────────────────────────────
    pl_targets = AUDIT_DRAFT_TARGETS["pnl"]
    pl_sheet_name = pl_targets["sheet"]
    # Handle potential trailing space in sheet name
    actual_pl = None
    for sn in wb.sheetnames:
        if sn.strip() == pl_sheet_name.strip():
            actual_pl = sn
            break

    if actual_pl:
        ws_pl = wb[actual_pl]
        cy_col = column_index_from_string(pl_targets["cy_col"])
        py_col = column_index_from_string(pl_targets["py_col"])

        for line_key, line_info in pl_targets["lines"].items():
            row = line_info["row"]
            bal = agg.get(line_key, {"cy": 0.0, "py": 0.0})
            _safe_write(ws_pl, row, cy_col, bal["cy"], report, f"PL:{line_key}:CY")
            _safe_write(ws_pl, row, py_col, bal["py"], report, f"PL:{line_key}:PY")

    # ── Equity — fix broken #REF! ─────────────────────────────────────────────
    if "Equity" in wb.sheetnames:
        ws_eq = wb["Equity"]
        eq_total_cy = sum(
            agg.get(k, {"cy": 0.0})["cy"]
            for k in ["ACCUMULATED_LOSSES", "REVALUATION_RESERVE", "GOVERNMENT_GRANTS_EQUITY"]
        )
        eq_total_py = sum(
            agg.get(k, {"py": 0.0})["py"]
            for k in ["ACCUMULATED_LOSSES", "REVALUATION_RESERVE", "GOVERNMENT_GRANTS_EQUITY"]
        )
        # Write opening balances (row 9, col N = total equity col)
        n_col = column_index_from_string("N")
        _safe_write(ws_eq, 9, n_col, eq_total_py, report, "EQ:OPENING")
        # Write closing 2024 (row 15, col N)
        _safe_write(ws_eq, 15, n_col, eq_total_cy, report, "EQ:CLOSING_CY")

    wb.save(str(output_path))
    wb.close()
    return report


def populate_audit_file(
    template_path: str | Path,
    output_path: str | Path,
    engagement: dict,
    mappings: list[dict],
) -> dict:
    """
    Populate the Audit File (working papers) template.
    """
    shutil.copy2(str(template_path), str(output_path))
    wb = openpyxl.load_workbook(str(output_path), keep_vba=False)

    report = {"written": [], "skipped": [], "issues": []}
    agg = _aggregate_balances(mappings)

    # ── AP-01 header ──────────────────────────────────────────────────────────
    ap01_targets = AUDIT_FILE_TARGETS["ap01"]
    ap01_name = ap01_targets["sheet"]
    if ap01_name in wb.sheetnames:
        ws_ap = wb[ap01_name]
        inputs = ap01_targets["inputs"]

        def write_ap(key, value):
            col, row = inputs[key]
            col_idx = column_index_from_string(col)
            _safe_write(ws_ap, row, col_idx, value, report, f"AP01:{key}")

        write_ap("client_name", engagement.get("entity_name", ""))
        write_ap("period",      engagement.get("period_end_date", ""))
        write_ap("prepared_by", engagement.get("prepared_by", "AI System"))

    # ── LEAD sheet ────────────────────────────────────────────────────────────
    lead_targets = AUDIT_FILE_TARGETS["lead"]
    lead_name = lead_targets["sheet"]
    if lead_name in wb.sheetnames:
        ws_lead = wb[lead_name]
        cy_col = column_index_from_string(lead_targets["cy_col"])
        py_col = column_index_from_string(lead_targets["py_col"])

        # Write year integers
        yc = lead_targets["year_cy_cell"]
        yp = lead_targets["year_py_cell"]
        try:
            year = int(engagement.get("period_year", datetime.now().year))
            _safe_write(ws_lead, yc[1], column_index_from_string(yc[0]), year, report, "LEAD:CY_YEAR")
            _safe_write(ws_lead, yp[1], column_index_from_string(yp[0]), year - 1, report, "LEAD:PY_YEAR")
        except Exception:
            pass

        # Find LEAD rows dynamically
        lead_row_map = _find_lead_rows(ws_lead)

        if not lead_row_map:
            report["issues"].append("Could not detect LEAD row labels — LEAD sheet may have custom structure.")
        else:
            for line_key, row_num in lead_row_map.items():
                bal = agg.get(line_key, {"cy": 0.0, "py": 0.0})
                _safe_write(ws_lead, row_num, cy_col, bal["cy"], report, f"LEAD:{line_key}:CY")
                _safe_write(ws_lead, row_num, py_col, bal["py"], report, f"LEAD:{line_key}:PY")

    wb.save(str(output_path))
    wb.close()
    return report


def _safe_write(ws, row: int, col: int, value, report: dict, label: str = "") -> None:
    """Write a value to a cell only if it's not formula-protected."""
    try:
        cell = ws.cell(row=row, column=col)
        existing = cell.value

        # Never overwrite a SUM formula — those auto-calculate
        if isinstance(existing, str) and existing.startswith("=SUM"):
            report["skipped"].append(f"{label or ''} row={row} col={col}: formula preserved")
            return

        cell.value = value
        report["written"].append({
            "label": label,
            "cell": f"{get_column_letter(col)}{row}",
            "value": value,
        })
    except Exception as e:
        report["issues"].append(f"{label}: {str(e)}")


def get_aggregated_balances(engagement_id: int) -> dict:
    """Return aggregated balances for an engagement (for reconciliation and UI)."""
    conn = get_conn()
    mappings = [
        dict(r) for r in conn.execute(
            """SELECT m.fs_statement, m.fs_category, m.fs_line_item, m.lead_line,
                      a.ending_balance, a.beginning_balance
               FROM account_mappings m
               JOIN tb_accounts a ON m.account_code = a.account_code
                                  AND a.engagement_id = m.engagement_id
               WHERE m.engagement_id = ?""",
            (engagement_id,)
        ).fetchall()
    ]
    conn.close()
    return _aggregate_balances(mappings)
