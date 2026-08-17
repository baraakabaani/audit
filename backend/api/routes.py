"""
FastAPI route definitions.

All arithmetic is deterministic Python — LLM is only used for classification suggestions.
"""

import json
import uuid
from datetime import datetime
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, HTTPException, UploadFile, File, Form, BackgroundTasks
from fastapi.responses import FileResponse

from backend.config import OUTPUTS_DIR
from backend.models.database import get_conn, init_db
from backend.models.schemas import (
    EngagementCreate, EngagementUpdate, MappingUpdate, BulkApprove, GenerateRequest
)
from backend.services.file_manager import save_upload, get_uploads, get_upload_path, register_output, get_output_path
from backend.services.tb_parser import parse_tb
from backend.services.classifier import classify_all_accounts, get_mappings, update_mapping, approve_mapping, bulk_approve
from backend.services.template_engine import analyze_template, populate_audit_draft, populate_audit_file, get_aggregated_balances
from backend.services.reconciliation import run_all_checks, save_validation_results

router = APIRouter()
init_db()


# ── Engagements ──────────────────────────────────────────────────────────────

@router.get("/engagements")
def list_engagements():
    conn = get_conn()
    rows = conn.execute("SELECT * FROM engagements ORDER BY id DESC").fetchall()
    conn.close()
    return [dict(r) for r in rows]


@router.post("/engagements")
def create_engagement(body: EngagementCreate):
    conn = get_conn()
    cur = conn.execute(
        """INSERT INTO engagements (name, entity_name, period, currency,
           overall_materiality, performance_materiality, trivial_threshold)
           VALUES (?,?,?,?,?,?,?)""",
        (body.name, body.entity_name, body.period, body.currency,
         body.overall_materiality, body.performance_materiality, body.trivial_threshold)
    )
    eid = cur.lastrowid
    conn.commit()
    conn.close()
    return {"id": eid, "name": body.name}


@router.get("/engagements/{eid}")
def get_engagement(eid: int):
    conn = get_conn()
    row = conn.execute("SELECT * FROM engagements WHERE id = ?", (eid,)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(404, "Engagement not found")
    return dict(row)


@router.patch("/engagements/{eid}")
def update_engagement(eid: int, body: EngagementUpdate):
    conn = get_conn()
    row = conn.execute("SELECT id FROM engagements WHERE id = ?", (eid,)).fetchone()
    if not row:
        conn.close()
        raise HTTPException(404, "Engagement not found")

    updates = {k: v for k, v in body.model_dump().items() if v is not None}
    if not updates:
        conn.close()
        return {"ok": True}

    sets = [f"{k} = ?" for k in updates]
    sets.append("updated_at = datetime('now')")
    vals = list(updates.values()) + [eid]
    conn.execute(f"UPDATE engagements SET {', '.join(sets)} WHERE id = ?", vals)
    conn.commit()
    conn.close()
    return {"ok": True}


# ── File Upload ───────────────────────────────────────────────────────────────

@router.post("/engagements/{eid}/upload")
async def upload_file(
    eid: int,
    file: UploadFile = File(...),
    file_type: str = Form(...),  # 'tb', 'prior_tb', 'audit_template', 'fs_template'
):
    conn = get_conn()
    row = conn.execute("SELECT id FROM engagements WHERE id = ?", (eid,)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(404, "Engagement not found")

    content = await file.read()
    result = save_upload(eid, file.filename, content, file_type)
    return result


@router.get("/engagements/{eid}/files")
def list_files(eid: int):
    return get_uploads(eid)


# ── TB Analysis ───────────────────────────────────────────────────────────────

@router.post("/engagements/{eid}/analyze-tb")
def analyze_tb(eid: int):
    """Parse the uploaded TB and store accounts in the database."""
    tb_path = get_upload_path(eid, "tb")
    if not tb_path:
        raise HTTPException(400, "No TB file uploaded. Please upload a Trial Balance first.")

    result = parse_tb(tb_path)

    # Persist to DB
    conn = get_conn()
    conn.execute("DELETE FROM tb_accounts WHERE engagement_id = ?", (eid,))

    for acc in result["accounts"]:
        conn.execute(
            """INSERT INTO tb_accounts
               (engagement_id, account_code, account_name, sub_account, account_type_raw,
                beginning_balance, period_activity, ending_balance, source_row,
                is_zero, is_unusual, unusual_reason)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                eid,
                acc["account_code"], acc["account_name"], acc.get("sub_account"),
                acc["account_type_raw"], acc["beginning_balance"],
                acc["period_activity"], acc["ending_balance"],
                acc["source_row"], int(acc["is_zero"]),
                int(acc["is_unusual"]), acc.get("unusual_reason"),
            )
        )

    # Update engagement metadata from TB
    meta = result.get("metadata", {})
    if meta.get("entity_name"):
        conn.execute(
            "UPDATE engagements SET entity_name = ?, updated_at = datetime('now') WHERE id = ? AND (entity_name IS NULL OR entity_name = '')",
            (meta["entity_name"], eid)
        )

    conn.commit()
    conn.close()

    return {
        "sheet_name": result["sheet_name"],
        "account_count": result["totals"]["account_count"],
        "total_ending": result["totals"]["total_ending"],
        "total_beginning": result["totals"]["total_beginning"],
        "balanced": result["balanced"],
        "by_type": result["totals"]["by_type"],
        "zero_count": result["totals"]["zero_count"],
        "unusual_count": result["totals"]["unusual_count"],
        "validation": result["validation"],
        "issues": result["issues"],
        "metadata": result["metadata"],
        "columns": result["columns"],
    }


@router.get("/engagements/{eid}/accounts")
def list_accounts(eid: int):
    conn = get_conn()
    rows = conn.execute(
        "SELECT * FROM tb_accounts WHERE engagement_id = ? ORDER BY account_code",
        (eid,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ── Template Analysis ─────────────────────────────────────────────────────────

@router.post("/engagements/{eid}/analyze-template")
def analyze_template_route(eid: int, template_type: str = "fs_template"):
    """Analyze an uploaded template workbook."""
    tpath = get_upload_path(eid, template_type)
    if not tpath:
        raise HTTPException(400, f"No {template_type} uploaded.")

    meta = analyze_template(tpath)

    # Store metadata
    conn = get_conn()
    conn.execute(
        "INSERT INTO template_metadata (engagement_id, file_type, metadata_json) VALUES (?,?,?)",
        (eid, template_type, json.dumps(meta))
    )
    conn.commit()
    conn.close()

    return meta


# ── Classification ────────────────────────────────────────────────────────────

@router.post("/engagements/{eid}/classify")
def classify_accounts(eid: int, use_ai: bool = True):
    """Run the classification engine on all TB accounts."""
    conn = get_conn()
    count = conn.execute(
        "SELECT COUNT(*) FROM tb_accounts WHERE engagement_id = ?", (eid,)
    ).fetchone()[0]
    conn.close()

    if count == 0:
        raise HTTPException(400, "No TB accounts found. Run analyze-tb first.")

    results = classify_all_accounts(eid, use_ai=use_ai)

    low = sum(1 for r in results if r["confidence_level"] == "LOW")
    medium = sum(1 for r in results if r["confidence_level"] == "MEDIUM")
    high = sum(1 for r in results if r["confidence_level"] == "HIGH")

    return {
        "total": len(results),
        "high_confidence": high,
        "medium_confidence": medium,
        "low_confidence": low,
        "needs_review": low + medium,
    }


@router.get("/engagements/{eid}/mappings")
def list_mappings(eid: int):
    return get_mappings(eid)


@router.patch("/engagements/{eid}/mappings/{account_code}")
def edit_mapping(eid: int, account_code: str, body: MappingUpdate):
    updates = {k: v for k, v in body.model_dump().items() if v is not None}
    if not updates:
        raise HTTPException(400, "No updates provided")
    update_mapping(eid, account_code, updates)
    return {"ok": True}


@router.post("/engagements/{eid}/mappings/{account_code}/approve")
def approve_account(eid: int, account_code: str):
    approve_mapping(eid, account_code)
    return {"ok": True}


@router.post("/engagements/{eid}/mappings/bulk-approve")
def bulk_approve_accounts(eid: int, body: BulkApprove):
    count = bulk_approve(eid, body.account_codes)
    return {"approved": count}


# ── Reconciliation / Validation ───────────────────────────────────────────────

@router.post("/engagements/{eid}/validate")
def validate_engagement(eid: int):
    results = run_all_checks(eid)
    save_validation_results(eid, results)

    errors = [r for r in results if r["result"] == "ERROR"]
    warnings = [r for r in results if r["result"] == "WARNING"]
    passes = [r for r in results if r["result"] == "PASS"]

    return {
        "total_checks": len(results),
        "errors": len(errors),
        "warnings": len(warnings),
        "passes": len(passes),
        "can_generate": len(errors) == 0,
        "results": results,
    }


@router.get("/engagements/{eid}/aggregated-balances")
def aggregated_balances(eid: int):
    return get_aggregated_balances(eid)


# ── Generate ──────────────────────────────────────────────────────────────────

@router.post("/engagements/{eid}/generate")
def generate_reports(eid: int, generate_audit_file: bool = True, generate_fs_draft: bool = True):
    """Generate the populated Excel output files."""
    conn = get_conn()
    engagement = conn.execute("SELECT * FROM engagements WHERE id = ?", (eid,)).fetchone()
    conn.close()

    if not engagement:
        raise HTTPException(404, "Engagement not found")

    eng = dict(engagement)
    entity = eng.get("entity_name", "Client")
    period = eng.get("period", "2024")

    # Build engagement context for template population
    eng_ctx = {
        "entity_name": entity,
        "period": period,
        "currency": eng.get("currency", "AED"),
        "location": "DUBAI - UNITED ARAB EMIRATES",
        "period_end_date": "31 December 2024",
        "period_end_words": "31 December 2024",
        "period_end_words_upper": "31 DECEMBER 2024",
        "period_start": "01 January 2024",
        "prior_period_end": "31 December 2023",
        "prior_period_words": "31 December 2023",
        "prior_period_start": "01 January 2023",
        "period_year": "2024",
        "authorised_person": "",
        "designation": "",
        "prepared_by": "AI System (Pending Auditor Review)",
    }

    mappings = get_mappings(eid)
    generated = []

    if generate_fs_draft:
        fs_path = get_upload_path(eid, "fs_template")
        if not fs_path:
            raise HTTPException(400, "No financial statement template uploaded.")

        out_name = f"{uuid.uuid4().hex}_FS_Draft.xlsx"
        out_path = OUTPUTS_DIR / out_name
        report = populate_audit_draft(fs_path, out_path, eng_ctx, mappings)
        register_output(eid, "fs_draft", out_name, fs_path.name)
        generated.append({
            "type": "fs_draft",
            "filename": f"{entity} - Financial Statements - {period}.xlsx",
            "stored": out_name,
            "cells_written": len(report["written"]),
            "issues": report["issues"],
        })

    if generate_audit_file:
        af_path = get_upload_path(eid, "audit_template")
        if not af_path:
            raise HTTPException(400, "No audit file template uploaded.")

        out_name = f"{uuid.uuid4().hex}_Audit_File.xlsx"
        out_path = OUTPUTS_DIR / out_name
        report = populate_audit_file(af_path, out_path, eng_ctx, mappings)
        register_output(eid, "audit_file", out_name, af_path.name)
        generated.append({
            "type": "audit_file",
            "filename": f"{entity} - Audit Report - {period}.xlsx",
            "stored": out_name,
            "cells_written": len(report["written"]),
            "issues": report["issues"],
        })

    return {"generated": generated}


# ── Download ──────────────────────────────────────────────────────────────────

@router.get("/engagements/{eid}/download/{file_type}")
def download_file(eid: int, file_type: str):
    out_path = get_output_path(eid, file_type)
    if not out_path:
        raise HTTPException(404, "File not yet generated.")

    conn = get_conn()
    eng = conn.execute("SELECT entity_name, period FROM engagements WHERE id = ?", (eid,)).fetchone()
    conn.close()

    entity = eng["entity_name"] if eng else "Client"
    period = eng["period"] if eng else "2024"

    names = {
        "fs_draft":   f"{entity} - Financial Statements - {period}.xlsx",
        "audit_file": f"{entity} - Audit Report - {period}.xlsx",
    }
    dl_name = names.get(file_type, out_path.name)

    return FileResponse(str(out_path), filename=dl_name, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.get("/engagements/{eid}/download-mapping-report")
def download_mapping_report(eid: int):
    """Generate and download the mapping report as Excel."""
    import openpyxl
    mappings = get_mappings(eid)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Account Mapping"

    headers = [
        "Account Code", "Account Name", "Account Type", "Ending Balance",
        "FS Statement", "FS Category", "FS Line Item", "Lead Line",
        "Confidence", "Confidence Level", "Source", "User Approved",
        "IFRS Reference", "Reason"
    ]
    ws.append(headers)

    for m in mappings:
        ws.append([
            m.get("account_code", ""),
            m.get("account_name", ""),
            m.get("account_type_raw", ""),
            m.get("ending_balance", 0),
            m.get("fs_statement", ""),
            m.get("fs_category", ""),
            m.get("fs_line_item", ""),
            m.get("lead_line", ""),
            m.get("confidence", 0),
            m.get("confidence_level", ""),
            m.get("source", ""),
            "Yes" if m.get("user_approved") else "No",
            m.get("ifrs_reference", ""),
            m.get("reason", ""),
        ])

    out_name = f"mapping_report_{eid}.xlsx"
    out_path = OUTPUTS_DIR / out_name
    wb.save(str(out_path))

    conn = get_conn()
    eng = conn.execute("SELECT entity_name, period FROM engagements WHERE id = ?", (eid,)).fetchone()
    conn.close()
    entity = eng["entity_name"] if eng else "Client"
    period = eng["period"] if eng else "2024"

    return FileResponse(
        str(out_path),
        filename=f"{entity} - Mapping Report - {period}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@router.get("/engagements/{eid}/download-audit-trail")
def download_audit_trail(eid: int):
    import openpyxl
    conn = get_conn()
    rows = conn.execute(
        "SELECT * FROM audit_trail WHERE engagement_id = ? ORDER BY created_at",
        (eid,)
    ).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Trail"
    ws.append(["Timestamp", "Action", "Account Code", "Field", "Old Value", "New Value", "Source", "Reason"])
    for r in rows:
        d = dict(r)
        ws.append([d.get("created_at"), d.get("action"), d.get("account_code"),
                   d.get("field"), d.get("old_value"), d.get("new_value"),
                   d.get("source"), d.get("reason")])

    out_name = f"audit_trail_{eid}.xlsx"
    out_path = OUTPUTS_DIR / out_name
    wb.save(str(out_path))
    return FileResponse(str(out_path), filename=f"Audit Trail - Engagement {eid}.xlsx",
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── Validation report download ─────────────────────────────────────────────────

@router.get("/engagements/{eid}/download-validation-report")
def download_validation_report(eid: int):
    import openpyxl
    conn = get_conn()
    rows = conn.execute(
        "SELECT * FROM validation_results WHERE engagement_id = ?", (eid,)
    ).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Validation Report"
    ws.append(["Check", "Result", "Expected", "Actual", "Difference", "Severity", "Explanation"])
    for r in rows:
        d = dict(r)
        ws.append([d.get("check_name"), d.get("result"), d.get("expected"),
                   d.get("actual"), d.get("difference"), d.get("severity"), d.get("explanation")])

    out_name = f"validation_report_{eid}.xlsx"
    out_path = OUTPUTS_DIR / out_name
    wb.save(str(out_path))
    return FileResponse(str(out_path), filename=f"Validation Report - Engagement {eid}.xlsx",
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
